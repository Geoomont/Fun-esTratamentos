from json import load
from openpyxl.utils import  get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl import load_workbook
from libs import *
from src.service.dataframe import *


def formata_excel(patch,nameFile):
    print('Iniciando formatação de Template')
    nome_arquivo=patch
    wb = load_workbook(nome_arquivo)
    ws = wb.active
    wl = ws.max_row
    wc = ws.max_column
    ws.title = 'Tratativas' #renomeando a sheet
    ws.sheet_view.showGridLines = False

    print('Iniciando Formatação')
        
        #Preenchimento
    my_fill = PatternFill(
            start_color='002060',
            fill_type='solid')

        #Borda
    my_bolder = Border(
            left= Side(None),
            right= Side(None))
        

        # Fonte
    my_font1 = Font(
                bold=True,
                color='ffffff',
                name='Verdana',
                size=9 )

    my_font2 = Font(
            bold=False,
            color='000000',
            name='Verdana',
                size=9 )                   

    alignment1 = Alignment(
                horizontal='center',
                vertical='center',
                text_rotation=0,
                wrap_text=False,
                shrink_to_fit=False,
                indent=0)

    alignment2 = Alignment(
                horizontal='right',
                vertical='center',
                text_rotation=0,
                wrap_text=False,
                shrink_to_fit=False,
                indent=0)
    
    '''formatando cabeçalho'''
    for coluna in range(1, wc+1):
        ws.column_dimensions[get_column_letter(coluna)].width =30

        for cell2 in range(1,2): 
            colformat = ws.cell(row = cell2, column = coluna)
            ws.row_dimensions[cell2].height = 12 
            colformat.border = my_bolder
            colformat.fill = my_fill
            colformat.font = my_font1

    '''Formatando linhas e colunas'''

    for coluna in range(1, wc+1):
        for cell2 in range(1, wl+1):
                        colformat = ws.cell(row = cell2, column = coluna)
                        ws.row_dimensions[cell2].height = 12        
                        colformat.border = my_bolder
                        # colformat.font = my_font3
                        colformat.alignment = alignment1
    print('finalizada a formatação das colunas e linhas')
    wb.save(nome_arquivo)
    wb.close()
    patch1 = patch #aqui
    nameFile1 = nameFile #aqui
    
    return responses.FileResponse(path=patch1, filename=nameFile1)